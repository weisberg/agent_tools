"""Tests for xli_companion.generate."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import polars as pl
import pytest

from xli_companion.generate import (
    generate_from_csv,
    generate_from_dataframe,
    generate_from_parquet,
    generate_multi_sheet,
)


@pytest.fixture()
def tmp_path_factory_dir(tmp_path: Path) -> Path:
    return tmp_path


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_generate_from_dataframe(tmp_path: Path) -> None:
    df = pl.DataFrame(
        {
            "name": ["Alice", "Bob"],
            "age": [30, 25],
            "active": [True, False],
        }
    )
    out = tmp_path / "out.xlsx"
    result = generate_from_dataframe(df, out)

    assert result == out
    assert out.exists()

    # Verify content with openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws is not None

    # Headers
    assert [cell.value for cell in ws[1]] == ["name", "age", "active"]
    # First data row
    assert ws.cell(2, 1).value == "Alice"
    assert ws.cell(2, 2).value == 30
    assert ws.cell(2, 3).value is True
    # Second data row
    assert ws.cell(3, 1).value == "Bob"
    assert ws.cell(3, 2).value == 25
    assert ws.cell(3, 3).value is False
    wb.close()


def test_generate_from_csv(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    csv_path.write_text("x,y\n1,hello\n2,world\n")

    xlsx_path = tmp_path / "from_csv.xlsx"
    result = generate_from_csv(csv_path, xlsx_path)

    assert result == xlsx_path
    assert xlsx_path.exists()

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    assert ws is not None
    assert [cell.value for cell in ws[1]] == ["x", "y"]
    assert ws.cell(2, 1).value == 1
    assert ws.cell(2, 2).value == "hello"
    assert ws.cell(3, 1).value == 2
    assert ws.cell(3, 2).value == "world"
    wb.close()


def test_generate_multi_sheet(tmp_path: Path) -> None:
    sheets = {
        "First": pl.DataFrame({"a": [1, 2]}),
        "Second": pl.DataFrame({"b": ["x", "y"]}),
    }
    out = tmp_path / "multi.xlsx"
    result = generate_multi_sheet(sheets, out)

    assert result == out
    assert out.exists()

    wb = openpyxl.load_workbook(out)
    assert "First" in wb.sheetnames
    assert "Second" in wb.sheetnames

    ws1 = wb["First"]
    assert ws1.cell(1, 1).value == "a"
    assert ws1.cell(2, 1).value == 1
    assert ws1.cell(3, 1).value == 2

    ws2 = wb["Second"]
    assert ws2.cell(1, 1).value == "b"
    assert ws2.cell(2, 1).value == "x"
    assert ws2.cell(3, 1).value == "y"
    wb.close()


def test_generate_from_parquet(tmp_path: Path) -> None:
    df = pl.DataFrame({"col1": [10, 20], "col2": ["foo", "bar"]})
    parquet_path = tmp_path / "input.parquet"
    df.write_parquet(parquet_path)

    xlsx_path = tmp_path / "from_parquet.xlsx"
    result = generate_from_parquet(parquet_path, xlsx_path)

    assert result == xlsx_path
    assert xlsx_path.exists()

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    assert ws is not None
    assert [cell.value for cell in ws[1]] == ["col1", "col2"]
    assert ws.cell(2, 1).value == 10
    assert ws.cell(2, 2).value == "foo"
    assert ws.cell(3, 1).value == 20
    assert ws.cell(3, 2).value == "bar"
    wb.close()
