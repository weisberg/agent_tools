"""Tests for OOXML artifact inspection checks."""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl

from xli_companion.checks.ooxml import (
    check_content_types,
    check_chart_relationships,
    check_shared_strings_integrity,
    check_vba_presence,
)


def _create_basic_xlsx(directory: Path) -> Path:
    """Create a minimal valid xlsx in the given directory."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    path = directory / "test.xlsx"
    wb.save(path)
    return path


def test_content_types_valid_workbook():
    """Create basic xlsx, verify no errors."""
    with tempfile.TemporaryDirectory() as tmpdir:
        path = _create_basic_xlsx(Path(tmpdir))
        findings = check_content_types(path)
        errors = [f for f in findings if f.severity.value == "error"]
        assert len(errors) == 0


def test_vba_detection():
    """Test with a regular xlsx (no VBA), verify no VBA finding."""
    with tempfile.TemporaryDirectory() as tmpdir:
        path = _create_basic_xlsx(Path(tmpdir))
        findings = check_vba_presence(path)
        assert len(findings) == 0


def test_shared_strings_integrity():
    """Create xlsx with string data, verify SST check passes."""
    with tempfile.TemporaryDirectory() as tmpdir:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "alpha"
        ws["A2"] = "beta"
        ws["A3"] = "gamma"
        path = Path(tmpdir) / "strings.xlsx"
        wb.save(path)

        findings = check_shared_strings_integrity(path)
        errors = [f for f in findings if f.code == "SST_COUNT_MISMATCH"]
        assert len(errors) == 0


def test_missing_file_returns_error():
    """Pass nonexistent path, verify graceful error."""
    fake_path = Path("/tmp/nonexistent_workbook_abc123.xlsx")
    for check_fn in [
        check_content_types,
        check_chart_relationships,
        check_shared_strings_integrity,
        check_vba_presence,
    ]:
        findings = check_fn(fake_path)
        assert len(findings) == 1
        assert findings[0].code == "FILE_NOT_FOUND"
        assert findings[0].severity.value == "error"
